"""
=============================
Author:Zangxiancheng


telephone:13359815071
Time:2022/5/11 
=============================
"""
"""
1.可以读取任意excel中的sheel表
2.写入任意excel中sheet表中的任意ceel中

"""
import openpyxl


class ExcelJar:

    def __init__(self, filename, sheetname):
        """
        :param filename: 文件路径
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取首行
        tittle = [i.value for i in res[0]]
        # 获取其他
        cases = []
        for item in res[1:]:
            r2 = [i.value for i in item]
            dic = dict(zip(tittle, r2))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        """
        :param row: 预输入行
        :param column: 预输入列
        :param value: 预输入值
        """
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)


if __name__ == '__main__':
    excel = ExcelJar(r'C:\Users\lenovo\Desktop\testcases.xlsx', '1')  # init 方法在创建对象的时候传入
    read = excel.read_data()
    write = excel.write_data(1, 2, 3)
    print(read)
    print(write)
    print("ceshi ")
